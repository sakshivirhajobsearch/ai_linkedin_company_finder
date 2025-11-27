import os, csv, json, threading, tkinter as tk
from tkinter import filedialog, messagebox
from concurrent.futures import ThreadPoolExecutor
import requests, socket, ssl, time
from datetime import datetime
from colorama import Fore, init
from openpyxl import Workbook
from openpyxl.styles import PatternFill

init(autoreset=True)

# ================= CONFIG =================
BG = "#000000"
FG = "#00FF41"
BTN = "#003300"
RED = "#FF3333"
GREEN = "#00FF41"

MAX_THREADS = 8
MAX_LENGTH = 40
BASE_OUTPUT_DIR = "output"

HEADERS = ["Website", "Domain", "DNS Working", "SSL Valid", "HTTP Status", "Final Status"]

VALID_TLDS = [
    "com","net","org","in","us","uk","de","fr","cn","jp","io",
    "gov","edu","biz","info","aero","nl","me","ai","res"
]

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# ================= UTILS =================
def fixed_width(text, width=MAX_LENGTH):
    if not text: text = ""
    return text[:width-3] + "..." if len(text) > width else text.ljust(width)

def is_valid_domain(text):
    if "." not in text: return False
    if text.lower().endswith((".png",".jpg",".jpeg",".gif",".html",".zip",".pdf",".txt",".xlsx")):
        return False
    return text.split(".")[-1] in VALID_TLDS

def normalize_url(domain):
    if domain.startswith("http"): return domain
    return "https://www." + domain if not domain.startswith("www.") else "https://" + domain

def extract_domain(url):
    return url.replace("https://","").replace("http://","").replace("www.","").split("/")[0]

def dns_check(domain):
    try:
        socket.setdefaulttimeout(5)
        socket.gethostbyname(domain)
        return True
    except: return False

def ssl_check(domain):
    try:
        ctx = ssl.create_default_context()
        with socket.create_connection((domain, 443), timeout=5) as sock:
            with ctx.wrap_socket(sock, server_hostname=domain):
                return True
    except: return False

def http_status(url):
    try:
        r = requests.get(url, timeout=7, headers={"User-Agent": "Mozilla/5.0"})
        return r.status_code
    except: return None

# ================= APP =================
class WebsiteChecker:
    def __init__(self, root):
        self.root = root
        self.root.title("> WEBSITE STATUS CHECKER")
        self.root.geometry("1200x850")
        self.root.configure(bg=BG)

        # ✅ GUI Ctrl Handling
        self.root.protocol("WM_DELETE_WINDOW", lambda: self.clean_shutdown("WINDOW CLOSED"))
        self.root.bind("<Control-c>", lambda e: self.clean_shutdown("CTRL+C GUI"))
        self.root.bind("<Break>", lambda e: self.clean_shutdown("CTRL+C GUI"))
        self.root.bind("<Escape>", lambda e: self.clean_shutdown("ESCAPE KEY"))

        self.websites = []
        self.pending = []
        self.total = 0
        self.done = 0
        self.valid = 0
        self.invalid = 0

        self.pause = False
        self.stop = False
        self.executor = None
        self.lock = threading.Lock()

        self.RUN_DIR = None
        self.STATE_FILE = None

        self.build_ui()

    # ================= UI =================
    def build_ui(self):
        top = tk.Frame(self.root, bg=BG)
        top.pack(fill=tk.X, padx=10, pady=5)

        self.input_entry = tk.Entry(top, width=60, bg="#101010", fg=FG)
        self.input_entry.pack(side=tk.LEFT, padx=5)

        tk.Button(top, text="BROWSE", command=self.browse, bg=BTN, fg=FG).pack(side=tk.LEFT)
        tk.Button(top, text="START", command=self.start, bg=BTN, fg=FG).pack(side=tk.LEFT, padx=10)
        tk.Button(top, text="PAUSE", command=self.pause_scan, bg=BTN, fg=FG).pack(side=tk.LEFT)
        tk.Button(top, text="RESUME", command=self.resume_scan, bg=BTN, fg=FG).pack(side=tk.LEFT, padx=10)
        tk.Button(top, text="CANCEL", command=lambda: self.clean_shutdown("CANCEL"), bg="#330000", fg=FG).pack(side=tk.LEFT)

        self.status_label = tk.Label(self.root, bg=BG, fg=FG, font=("Consolas", 11, "bold"))
        self.status_label.pack(pady=5)

        self.progress_canvas = tk.Canvas(self.root, height=22, bg="#101010")
        self.progress_canvas.pack(fill=tk.X)

        self.log_box = tk.Text(self.root, bg="#050505", fg=FG, font=("Consolas", 10))
        self.log_box.pack(fill=tk.BOTH, expand=True)

        self.log_box.tag_config("valid", foreground=GREEN)
        self.log_box.tag_config("invalid", foreground=RED)

        header = f"{'WEBSITE':<40} | {'DOMAIN':<40} | DNS | SSL | HTTP | STATUS\n"
        self.log_box.insert(tk.END, header)

    def browse(self):
        file = filedialog.askopenfilename()
        if file:
            self.input_entry.delete(0, tk.END)
            self.input_entry.insert(0, file)

    # ================= OUTPUT SYSTEM =================
    def create_run_folder(self):
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        self.RUN_DIR = os.path.join(BASE_OUTPUT_DIR, timestamp)
        os.makedirs(self.RUN_DIR, exist_ok=True)

        self.STATE_FILE = os.path.join(self.RUN_DIR, "autosave.json")

        self.ALL_CSV = os.path.join(self.RUN_DIR, "all_results.csv")
        self.VALID_CSV = os.path.join(self.RUN_DIR, "valid_websites.csv")
        self.INVALID_CSV = os.path.join(self.RUN_DIR, "invalid_websites.csv")
        self.VALID_TXT = os.path.join(self.RUN_DIR, "valid_websites.txt")
        self.INVALID_TXT = os.path.join(self.RUN_DIR, "invalid_websites.txt")
        self.VALID_XLSX = os.path.join(self.RUN_DIR, "valid_websites.xlsx")
        self.INVALID_XLSX = os.path.join(self.RUN_DIR, "invalid_websites.xlsx")

        self.valid_wb = Workbook()
        self.invalid_wb = Workbook()
        self.valid_ws = self.valid_wb.active
        self.invalid_ws = self.invalid_wb.active
        self.valid_ws.append(HEADERS)
        self.invalid_ws.append(HEADERS)

        for f in [self.ALL_CSV, self.VALID_CSV, self.INVALID_CSV]:
            with open(f,"w",newline="",encoding="utf-8") as file:
                csv.writer(file).writerow(HEADERS)

        with open(self.VALID_TXT,"w") as f: f.write("Website\n")
        with open(self.INVALID_TXT,"w") as f: f.write("Website\n")

    # ================= WORK =================
    def start(self):
        self.create_run_folder()

        path = self.input_entry.get()
        if not os.path.exists(path):
            messagebox.showerror("Error", "Input file not found")
            return

        with open(path) as f:
            self.websites = [x.strip() for x in f if x.strip()]

        self.pending = list(self.websites)
        self.total = len(self.websites)

        self.executor = ThreadPoolExecutor(max_workers=MAX_THREADS)

        for site in self.pending:
            self.executor.submit(self.process_site, site)

    def process_site(self, site):
        while self.pause and not self.stop:
            time.sleep(0.2)
        if self.stop: return

        if not is_valid_domain(site):
            self.store_result(site, "", False, False, None, "INVALID")
            return

        website = normalize_url(site)
        domain = extract_domain(website)

        dns = dns_check(domain)
        ssl_ok = ssl_check(domain)
        http = http_status(website)

        status = "VALID" if dns and ssl_ok and http and http < 400 else "INVALID"
        self.store_result(website, domain, dns, ssl_ok, http, status)

    # ================= STORAGE =================
    def store_result(self, website, domain, dns, ssl_ok, http, final_status):
        row = [website, domain, dns, ssl_ok, http, final_status]

        with self.lock:
            for target, ws, fill in [
                (self.VALID_CSV, self.valid_ws, GREEN_FILL) if final_status == "VALID"
                else (self.INVALID_CSV, self.invalid_ws, RED_FILL)
            ]:
                with open(target,"a",newline="",encoding="utf-8") as f:
                    csv.writer(f).writerow(row)

                ws.append(row)
                for col in range(1, 7):
                    ws.cell(row=ws.max_row, column=col).fill = fill

                if final_status=="VALID":
                    self.valid_wb.save(self.VALID_XLSX)
                else:
                    self.invalid_wb.save(self.INVALID_XLSX)

            with open(self.ALL_CSV,"a",newline="",encoding="utf-8") as f:
                csv.writer(f).writerow(row)

            with open(self.VALID_TXT if final_status=="VALID" else self.INVALID_TXT,"a") as f:
                f.write(website+"\n")

        console_line = f"{fixed_width(website)} | {fixed_width(domain)} | {dns} | {ssl_ok} | {http} | {final_status}"
        print((Fore.GREEN if final_status=="VALID" else Fore.RED)+console_line)

        self.root.after(0, self.display_gui, console_line, final_status)

        self.done += 1
        if final_status=="VALID": self.valid+=1
        else: self.invalid+=1

        self.update_progress()

    def display_gui(self, text, status):
        self.log_box.insert(tk.END, text+"\n", "valid" if status=="VALID" else "invalid")
        self.log_box.see(tk.END)

    # ================= STATE =================
    def pause_scan(self):
        self.pause = True
        self.save_state("PAUSED")

    def resume_scan(self):
        self.pause = False
        self.restore_state()

    def save_state(self, reason):
        with open(self.STATE_FILE,"w") as f:
            json.dump({
                "reason":reason,
                "done":self.done,
                "remaining":self.websites[self.done:]
            }, f, indent=4)

    def restore_state(self):
        if not os.path.exists(self.STATE_FILE): return
        with open(self.STATE_FILE) as f:
            state=json.load(f)
        for site in state["remaining"]:
            self.executor.submit(self.process_site, site)

    # ================= EXIT =================
    def update_progress(self):
        percent = (self.done/self.total)*100
        self.status_label.config(
            text=f"Progress: {percent:.2f}% | Done:{self.done} | Valid:{self.valid} | Invalid:{self.invalid}"
        )

    def clean_shutdown(self, reason):
        print(f"\n[CLEAN EXIT] {reason}")
        self.save_state(reason)
        self.stop = True
        if self.executor:
            self.executor.shutdown(wait=False)
        self.root.destroy()
        os._exit(0)

# ================= MAIN =================
if __name__ == "__main__":
    root = tk.Tk()
    app = WebsiteChecker(root)
    try:
        root.mainloop()
    except KeyboardInterrupt:
        print("\n[CTRL+C CAUGHT IN MAINLOOP]")
        app.clean_shutdown("CTRL+C TERMINAL")
